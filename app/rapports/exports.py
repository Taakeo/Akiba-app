import io

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from ..pdf_utils import logo_flowable

# NOTE : gabarit générique (date / regroupement / quantité / montant). Le
# modèle comptable exact utilisé par Akiba (§11.3 spec) n'est pas disponible
# dans ce dépôt — à ajuster une fois le fichier de référence du comptable
# fourni, pour que l'export s'insère directement dans son classeur existant.


def _entete(ws, titre, date_debut, date_fin):
    ws["A1"] = "AKIBA APP"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = titre
    ws["A3"] = f"Période : {date_debut.isoformat()} au {date_fin.isoformat()}"


def export_ventes_xlsx(lignes, total, date_debut, date_fin, group_by):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventes"
    _entete(ws, "Rapport des ventes", date_debut, date_fin)

    ws.append([])
    ws.append([group_by.capitalize(), "Quantité", "Montant"])
    for cell in ws[5]:
        cell.font = Font(bold=True)

    for ligne in lignes:
        ws.append([ligne["libelle"], ligne["quantite"], ligne["total"]])

    ws.append(["TOTAL", "", total])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15

    return _vers_bytesio(wb)


def export_achats_xlsx(lignes, total, date_debut, date_fin, group_by):
    wb = Workbook()
    ws = wb.active
    ws.title = "Achats"
    _entete(ws, "Rapport des achats", date_debut, date_fin)

    ws.append([])
    ws.append([group_by.capitalize(), "Nombre d'achats", "Montant"])
    for cell in ws[5]:
        cell.font = Font(bold=True)

    for ligne in lignes:
        ws.append([ligne["libelle"], ligne["nombre"], ligne["total"]])

    ws.append(["TOTAL", "", total])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    return _vers_bytesio(wb)


def _vers_bytesio(wb):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _table_pdf(titre, date_debut, date_fin, entetes, lignes_texte, total_texte):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()

    logo = logo_flowable(largeur_mm=30)
    elements = [logo] if logo is not None else [Paragraph("AKIBA APP", styles["Title"])]
    elements += [
        Paragraph(titre, styles["Heading2"]),
        Paragraph(f"Période : {date_debut.isoformat()} au {date_fin.isoformat()}", styles["Normal"]),
        Spacer(1, 16),
    ]

    data = [entetes, *lignes_texte, total_texte]
    table = Table(data, hAlign="LEFT", colWidths=[220, 120, 120][: len(entetes)])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0e7df")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.grey),
                ("LINEABOVE", (0, -1), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_ventes_pdf(lignes, total, date_debut, date_fin, group_by):
    lignes_texte = [[l["libelle"], str(l["quantite"]), f"{l['total']:,}".replace(",", " ")] for l in lignes]
    total_texte = ["TOTAL", "", f"{total:,}".replace(",", " ")]
    return _table_pdf(
        "Rapport des ventes",
        date_debut,
        date_fin,
        [group_by.capitalize(), "Quantité", "Montant"],
        lignes_texte,
        total_texte,
    )


def export_achats_pdf(lignes, total, date_debut, date_fin, group_by):
    lignes_texte = [[l["libelle"], str(l["nombre"]), f"{l['total']:,}".replace(",", " ")] for l in lignes]
    total_texte = ["TOTAL", "", f"{total:,}".replace(",", " ")]
    return _table_pdf(
        "Rapport des achats",
        date_debut,
        date_fin,
        [group_by.capitalize(), "Nombre d'achats", "Montant"],
        lignes_texte,
        total_texte,
    )
