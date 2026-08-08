"""Import en masse (produits, salariés) depuis un fichier Excel (§9 amélioration).

Volontairement strict et non-destructif : une ligne dont une référence
(poste, catégorie...) ne correspond à rien d'existant est rejetée avec une
explication précise plutôt que de créer une fiche à moitié remplie ou une
nouvelle catégorie fantôme par erreur de frappe. Une fiche déjà présente
(même nom, insensible à la casse) est ignorée plutôt qu'écrasée — un import
répété par erreur ne doit jamais dupliquer ni modifier silencieusement les
données existantes."""

import io
from datetime import date, datetime

from openpyxl import Workbook, load_workbook

from ..extensions import db
from ..models import (
    Categorie,
    FREQUENCES_REMUNERATION,
    Fournisseur,
    ParametresRH,
    Poste,
    PrixProduit,
    Produit,
    Projet,
    Salarie,
    SousCategorie,
    TYPES_CONTRAT,
    TypeTarif,
)

ENTETES_PRODUITS = [
    "Nom du produit",
    "Poste",
    "Catégorie",
    "Sous-catégorie",
    "Fournisseur principal",
    "Unité",
    "Prix d'achat",
    "Prix de référence",
    "Stock actuel",
    "Seuil d'alerte",
    "Code-barres",
]

EXEMPLE_PRODUIT = [
    "Riz blanc 1kg",
    "Boutique",
    "Alimentation",
    "",
    "",
    "kg",
    2000,
    2500,
    50,
    10,
    "",
]

ENTETES_SALARIES = [
    "Nom complet",
    "Téléphone",
    "Fonction",
    "Type de contrat",
    "Date d'embauche (JJ/MM/AAAA)",
    "Poste",
    "Projet",
    "Salaire habituel",
    "Fréquence de versement",
    "Quota de congés payés (jours/an)",
]

EXEMPLE_SALARIE = [
    "Rabe Andry",
    "034 00 000 00",
    "Vendeur",
    "CDI",
    "01/03/2024",
    "Boutique",
    "",
    150000,
    "Mensuel",
    "",
]


def _generer_modele(titre_feuille, entetes, exemple):
    wb = Workbook()
    ws = wb.active
    ws.title = titre_feuille
    ws.append(entetes)
    ws.append(exemple)
    for col_idx in range(1, len(entetes) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 24

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generer_modele_xlsx():
    """Classeur vierge (en-têtes + une ligne d'exemple) pour l'import de produits."""
    return _generer_modele("Produits", ENTETES_PRODUITS, EXEMPLE_PRODUIT)


def generer_modele_salaries_xlsx():
    """Classeur vierge (en-têtes + une ligne d'exemple) pour l'import de salariés."""
    return _generer_modele("Salariés", ENTETES_SALARIES, EXEMPLE_SALARIE)


def _valeur(cell):
    if cell is None:
        return ""
    return str(cell).strip()


def _entier(cell):
    if cell is None or str(cell).strip() == "":
        return None
    try:
        return int(float(cell))
    except (TypeError, ValueError):
        return "invalide"


def importer_produits_excel(fichier):
    """Lit le classeur, crée les produits valides, journalise le reste.
    Retourne {"crees": [...], "ignores": [...], "erreurs": [(ligne, message)]}."""
    try:
        wb = load_workbook(fichier, data_only=True)
    except Exception as exc:  # noqa: BLE001 - on veut un message utilisateur, pas une trace
        return {"crees": [], "ignores": [], "erreurs": [(0, f"Fichier illisible : {exc}")]}

    ws = wb.active

    postes = {p.name.strip().lower(): p for p in Poste.query.filter_by(is_archived=False)}
    # Une Catégorie est désormais rattachée à un seul Poste (deux postes
    # différents peuvent avoir une catégorie de même nom) : on indexe donc
    # par (poste_id, nom) plutôt que par nom seul.
    categories_par_poste_id = {}
    for c in Categorie.query.filter_by(is_archived=False):
        categories_par_poste_id.setdefault(c.poste_id, {})[c.name.strip().lower()] = c
    fournisseurs = {f.name.strip().lower(): f for f in Fournisseur.query.filter_by(is_archived=False)}
    noms_existants = {p.name.strip().lower() for p in Produit.query}
    # Tarif par défaut (typiquement "Standard") : le prix de référence importé
    # y devient directement un prix de vente manuel, sinon un produit importé
    # avec un "Prix de référence" resterait invendable au PDV tant qu'aucun
    # rabais automatique n'est configuré sur ce tarif (voir Produit.prix_pour).
    tarif_par_defaut = TypeTarif.query.filter_by(is_default=True, is_archived=False).first()

    resultat = {"crees": [], "ignores": [], "erreurs": []}

    lignes = list(ws.iter_rows(min_row=2, values_only=True))
    for i, ligne in enumerate(lignes, start=2):
        if ligne is None or all(c is None or str(c).strip() == "" for c in ligne):
            continue  # ligne vide : ignorée silencieusement, pas une erreur

        valeurs = list(ligne) + [None] * (len(ENTETES_PRODUITS) - len(ligne))
        nom = _valeur(valeurs[0])
        poste_nom = _valeur(valeurs[1])
        categorie_nom = _valeur(valeurs[2])
        sous_categorie_nom = _valeur(valeurs[3])
        fournisseur_nom = _valeur(valeurs[4])
        unite = _valeur(valeurs[5]) or "unité"
        prix_achat = _entier(valeurs[6])
        prix_reference = _entier(valeurs[7])
        stock_quantite = _entier(valeurs[8])
        seuil_alerte = _entier(valeurs[9])
        code_barres = _valeur(valeurs[10]) or None

        if not nom:
            resultat["erreurs"].append((i, "Nom du produit manquant."))
            continue

        if nom.strip().lower() in noms_existants:
            resultat["ignores"].append((i, f"« {nom} » existe déjà — ligne ignorée."))
            continue

        poste = postes.get(poste_nom.strip().lower())
        if not poste:
            resultat["erreurs"].append((i, f"Poste « {poste_nom} » introuvable (créez-le d'abord dans Administration)."))
            continue

        categorie = categories_par_poste_id.get(poste.id, {}).get(categorie_nom.strip().lower())
        if not categorie:
            resultat["erreurs"].append(
                (i, f"Catégorie « {categorie_nom} » introuvable sous le poste « {poste_nom} » (créez-la d'abord dans Administration).")
            )
            continue

        sous_categorie = None
        if sous_categorie_nom:
            sous_categorie = SousCategorie.query.filter_by(
                categorie_id=categorie.id, is_archived=False
            ).filter(db.func.lower(SousCategorie.name) == sous_categorie_nom.lower()).first()
            if not sous_categorie:
                resultat["erreurs"].append(
                    (i, f"Sous-catégorie « {sous_categorie_nom} » introuvable sous « {categorie_nom} ».")
                )
                continue

        fournisseur = None
        if fournisseur_nom:
            fournisseur = fournisseurs.get(fournisseur_nom.strip().lower())
            if not fournisseur:
                resultat["erreurs"].append((i, f"Fournisseur « {fournisseur_nom} » introuvable."))
                continue

        champ_invalide = None
        for champ, valeur in (
            ("Prix d'achat", prix_achat),
            ("Prix de référence", prix_reference),
            ("Stock actuel", stock_quantite),
            ("Seuil d'alerte", seuil_alerte),
        ):
            if valeur == "invalide":
                champ_invalide = champ
                break
        if champ_invalide:
            resultat["erreurs"].append((i, f"« {champ_invalide} » doit être un nombre entier."))
            continue

        produit = Produit(
            name=nom,
            poste_id=poste.id,
            categorie_id=categorie.id,
            sous_categorie_id=sous_categorie.id if sous_categorie else None,
            fournisseur_principal_id=fournisseur.id if fournisseur else None,
            unite=unite,
            prix_achat=prix_achat,
            prix_reference=prix_reference,
            stock_quantite=stock_quantite or 0,
            seuil_alerte=seuil_alerte,
            code_barres=code_barres,
        )
        db.session.add(produit)
        if prix_reference and tarif_par_defaut:
            db.session.flush()  # attribue produit.id, requis par PrixProduit
            db.session.add(
                PrixProduit(produit_id=produit.id, type_tarif_id=tarif_par_defaut.id, montant=prix_reference)
            )
        noms_existants.add(nom.strip().lower())
        resultat["crees"].append((i, nom))

    if resultat["crees"]:
        db.session.commit()
    else:
        db.session.rollback()

    return resultat


def _date(cell):
    if cell is None or str(cell).strip() == "":
        return None
    if isinstance(cell, datetime):
        return cell.date()
    if isinstance(cell, date):
        return cell
    try:
        return datetime.strptime(str(cell).strip(), "%d/%m/%Y").date()
    except ValueError:
        return "invalide"


def importer_salaries_excel(fichier):
    """Même logique que l'import produits (voir docstring du module),
    appliquée aux fiches salarié. Retourne la même structure de résultat."""
    try:
        wb = load_workbook(fichier, data_only=True)
    except Exception as exc:  # noqa: BLE001 - message utilisateur, pas une trace
        return {"crees": [], "ignores": [], "erreurs": [(0, f"Fichier illisible : {exc}")]}

    ws = wb.active

    postes = {p.name.strip().lower(): p for p in Poste.query.filter_by(is_archived=False)}
    projets = {p.name.strip().lower(): p for p in Projet.query.filter_by(is_archived=False)}
    contrats_valides = {c.lower(): c for c in TYPES_CONTRAT}
    frequences_valides = {label.lower(): code for code, label in FREQUENCES_REMUNERATION}
    noms_existants = {s.nom.strip().lower() for s in Salarie.query}
    quota_defaut = ParametresRH.get().quota_conges_defaut

    resultat = {"crees": [], "ignores": [], "erreurs": []}

    lignes = list(ws.iter_rows(min_row=2, values_only=True))
    for i, ligne in enumerate(lignes, start=2):
        if ligne is None or all(c is None or str(c).strip() == "" for c in ligne):
            continue

        valeurs = list(ligne) + [None] * (len(ENTETES_SALARIES) - len(ligne))
        nom = _valeur(valeurs[0])
        telephone = _valeur(valeurs[1]) or None
        fonction = _valeur(valeurs[2]) or None
        contrat_saisi = _valeur(valeurs[3])
        date_embauche = _date(valeurs[4])
        poste_nom = _valeur(valeurs[5])
        projet_nom = _valeur(valeurs[6])
        salaire_habituel = _entier(valeurs[7])
        frequence_saisie = _valeur(valeurs[8])
        quota_conges = _entier(valeurs[9])

        if not nom:
            resultat["erreurs"].append((i, "Nom complet manquant."))
            continue

        if nom.strip().lower() in noms_existants:
            resultat["ignores"].append((i, f"« {nom} » existe déjà — ligne ignorée."))
            continue

        type_contrat = None
        if contrat_saisi:
            type_contrat = contrats_valides.get(contrat_saisi.strip().lower())
            if not type_contrat:
                resultat["erreurs"].append(
                    (i, f"Type de contrat « {contrat_saisi} » inconnu (valeurs possibles : {', '.join(TYPES_CONTRAT)}).")
                )
                continue

        if date_embauche == "invalide":
            resultat["erreurs"].append((i, "Date d'embauche invalide (format attendu JJ/MM/AAAA)."))
            continue

        poste = None
        if poste_nom:
            poste = postes.get(poste_nom.strip().lower())
            if not poste:
                resultat["erreurs"].append((i, f"Poste « {poste_nom} » introuvable (créez-le d'abord dans Administration)."))
                continue

        projet = None
        if projet_nom:
            projet = projets.get(projet_nom.strip().lower())
            if not projet:
                resultat["erreurs"].append((i, f"Projet « {projet_nom} » introuvable (créez-le d'abord dans Administration)."))
                continue

        frequence_remuneration = None
        if frequence_saisie:
            frequence_remuneration = frequences_valides.get(frequence_saisie.strip().lower())
            if not frequence_remuneration:
                labels = ", ".join(label for _, label in FREQUENCES_REMUNERATION)
                resultat["erreurs"].append(
                    (i, f"Fréquence de versement « {frequence_saisie} » inconnue (valeurs possibles : {labels}).")
                )
                continue

        champ_invalide = None
        for champ, valeur in (("Salaire habituel", salaire_habituel), ("Quota de congés payés", quota_conges)):
            if valeur == "invalide":
                champ_invalide = champ
                break
        if champ_invalide:
            resultat["erreurs"].append((i, f"« {champ_invalide} » doit être un nombre entier."))
            continue

        salarie = Salarie(
            nom=nom,
            telephone=telephone,
            fonction=fonction,
            type_contrat=type_contrat,
            date_embauche=date_embauche,
            poste_id=poste.id if poste else None,
            projet_id=projet.id if projet else None,
            salaire_habituel=salaire_habituel,
            frequence_remuneration=frequence_remuneration,
            quota_conges=quota_conges if quota_conges is not None else quota_defaut,
        )
        db.session.add(salarie)
        noms_existants.add(nom.strip().lower())
        resultat["crees"].append((i, nom))

    if resultat["crees"]:
        db.session.commit()
    else:
        db.session.rollback()

    return resultat
